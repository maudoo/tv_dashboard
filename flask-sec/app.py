from flask import Flask, jsonify
from flask_cors import CORS
import dropbox
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, timedelta
import requests
import json
import os

app = Flask(__name__)
CORS(app)

def getDropbox():
    APP_KEY = os.getenv("DROPBOX_APP_KEY")
    APP_SECRET = os.getenv("DROPBOX_APP_SECRET")
    REFRESH_TOKEN = os.getenv("DROPBOX_REFRESH_TOKEN")

    if not all([APP_KEY, APP_SECRET, REFRESH_TOKEN]):
        raise RuntimeError("Missing Dropbox environment variables")

    return dropbox.Dropbox(
        oauth2_refresh_token=REFRESH_TOKEN,
        app_key=APP_KEY,
        app_secret=APP_SECRET
    )

@app.route('/')
def hello():
    return "Developed by Ahmad!"


@app.route('/LoadImages')
def LoadImages():
    """
    The function `LoadImages` retrieves web links to images hosted on Dropbox from a specified folder
    path.
    :return: The `LoadImages` function returns a list of web links to the images hosted on Dropbox after
    processing the files in the specified folder path.
    """

    with open("./data/MetaData.json", 'r') as file:
        IMAGES_FILE_PATH = json.load(file)["Images_Filepath"] # Path to folder that contains slideshow pictures in DROPBOX  

    dbx = getDropbox() # Get dropbox object to interact with files in dropbox account
    links = []  #Will contain the web links to the images hosted on dropbox
    entries = dbx.files_list_folder(IMAGES_FILE_PATH).entries
    for entry in entries:
        if isinstance(entry, dropbox.files.FileMetadata):
            try:
                # Try to get existing shared links for the file
                shared_links = dbx.sharing_list_shared_links(path=entry.path_lower).links
                # Check if there are any existing shared links
                if shared_links:
                    link = shared_links[0].url.replace("dl=0", "raw=1")  # Use the first existing link
                else:
                    # No existing links, create a new one
                    shared_link_metadata = dbx.sharing_create_shared_link_with_settings(entry.path_lower)
                    link = shared_link_metadata.url.replace("dl=0", "raw=1")
                links.append(link)
            
            except dropbox.exceptions.ApiError as api_err:
                print(f"LoadImages: Error obtaining link for {entry.path_lower}: {api_err}")
    return links

@app.route('/Iqamahs')
def Iqamahs():
    """
    The function `Iqamahs` checks if a file containing prayer times has been modified, downloads and
    processes the data if needed, and returns the prayer times in a JSON format.
    :return: The function `Iqamahs()` returns a JSON response containing the iqamah times for different
    prayers, converted to 12-hour format if they are in 24-hour format.
    """
    dbx = getDropbox()  # Get dropbox object to interact with files in dropbox account

    # Retrieve path to folder that contains Iqamahs excel file in DROPBOX
    with open("./data/MetaData.json", 'r') as file:
        IQAMAHS_FILE_PATH = json.load(file)["Iqamahs_Filepath"]

    # Get when what the last time the Iqamahs file was modified
    metadata = dbx.files_get_metadata(IQAMAHS_FILE_PATH)
    modified_date = metadata.server_modified

    # Get the saved modified date for the iqamahs file
    with open("./data/Dropbox_Last_Modified.json", 'r') as json_file:
        dropbox_json = json.load(json_file)
        saved_modified_date = dropbox_json['Iqamahs']

    # Compare the dates. If they do not equal then the file was modified
    if saved_modified_date != str(modified_date):
        # Store the new modified date of the iqamahs file. This will be the new date checked in the future to see if the file was changed again
        with open("./data/Dropbox_Last_Modified.json", 'w') as json_file:
            dropbox_json['Iqamahs'] = str(modified_date)
            json.dump(dropbox_json, json_file)

        # Download the file
        _, res = dbx.files_download(IQAMAHS_FILE_PATH)
        
        # Load the workbook
        workbook = load_workbook(filename=BytesIO(res.content))
        
        # Assuming your data is in the first sheet
        first_sheet_name = workbook.sheetnames[0]
        sheet = workbook[first_sheet_name]
        
        # Construct json object for iqamahs data
        iqamahs = {
            "Fajr": sheet['B2'].value,
            "Dhuhr": sheet['B3'].value,
            "Asr": sheet['B4'].value,
            "Maghrib": sheet['B5'].value,
            "Isha": sheet['B6'].value,
            "Jummah Khutbah": sheet['B7'].value,
            "Jummah Iqamah": sheet['B8'].value
        }
        
        # Save the new iqamahs data
        with open("./data/Iqamahs_Save.json", 'w') as json_file:
            json.dump(iqamahs, json_file)

    else:
        #Dates match meaning the file was not changed. Load up old data
        with open("./data/Iqamahs_Save.json", 'r') as json_file:
            iqamahs = json.load(json_file)

    for key in iqamahs.keys():
        # Any prayers that do not have an Iqamah assigned, convert to blank string ""
        if not iqamahs[key]:
            iqamahs[key] = ""
        
    return jsonify(iqamahs)
#Initialize Flask and run on port 7000
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 7000))  # default for local testing
    app.run(host="0.0.0.0", port=port)
